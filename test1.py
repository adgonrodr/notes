from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook


def normalise_merged_cells(ws) -> None:
    """Normalise merged cells by copying the anchor value across the merged range.

    Args:
        ws: An openpyxl worksheet.

    Returns:
        None. The worksheet is modified in-place.
    """
    for merged_range in list(ws.merged_cells.ranges):
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        if value is None:
            continue
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                ws.cell(r, c).value = value


def sheet_to_grid(ws) -> List[List[object]]:
    """Convert a worksheet to a 2D grid of cell values.

    Whitespace-only strings are normalised to None.

    Args:
        ws: An openpyxl worksheet.

    Returns:
        A list of rows, where each row is a list of cell values.
    """
    grid: List[List[object]] = []
    for r in range(1, ws.max_row + 1):
        row: List[object] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and not v.strip():
                v = None
            row.append(v)
        grid.append(row)
    return grid


def _ffill(values: List[object]) -> List[object]:
    """Forward-fill None values within a single list.

    Args:
        values: A list of values which may include None entries.

    Returns:
        A new list where None entries are replaced with the most recent non-None value.
    """
    out: List[object] = []
    last: Optional[object] = None
    for v in values:
        if v is None:
            out.append(last)
        else:
            last = v
            out.append(v)
    return out


def build_flat_columns(
    header_rows: Sequence[Sequence[object]],
    sep: str = " | ",
    drop_repeated: bool = True,
) -> List[str]:
    """Flatten 1..N header rows into a single list of column names.

    Approach:
      - Forward-fill blanks in each header row (useful for grouped headers from merged cells).
      - Join non-empty parts per column using a separator.
      - Ensure uniqueness (duplicate names get a suffix like " (2)").

    Args:
        header_rows: A sequence of header rows. Each header row is a sequence of cell values.
        sep: Separator used to join header parts.
        drop_repeated: If True, consecutive identical parts are collapsed.

    Returns:
        A list of flattened, unique column names.
    """
    if not header_rows:
        return []

    width = max(len(r) for r in header_rows)
    norm = [list(r) + [None] * (width - len(r)) for r in header_rows]
    norm = [_ffill(r) for r in norm]

    cols: List[str] = []
    for j in range(width):
        parts: List[str] = []
        prev: Optional[str] = None
        for i in range(len(norm)):
            v = norm[i][j]
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if drop_repeated and prev == s:
                continue
            parts.append(s)
            prev = s
        cols.append(sep.join(parts) if parts else f"col_{j}")

    seen: Dict[str, int] = {}
    unique_cols: List[str] = []
    for c in cols:
        n = seen.get(c, 0)
        unique_cols.append(c if n == 0 else f"{c} ({n + 1})")
        seen[c] = n + 1

    return unique_cols


def last_data_row_in_span(
    grid: List[List[object]],
    start_row: int,
    col_span: Tuple[int, int],
    max_blank_streak: int = 1,
) -> int:
    """Infer the last row of a table block based on data presence.

    Scans downward starting at start_row, looking for any non-empty cell
    in the given column span. Once max_blank_streak consecutive blank rows
    are seen, scanning stops and the most recent row containing data is returned.

    Args:
        grid: 2D grid of worksheet values.
        start_row: 1-based row index to start scanning.
        col_span: A (left_col, right_col) 1-based inclusive column span.
        max_blank_streak: Number of consecutive blank rows that ends the scan.

    Returns:
        The 1-based index of the last row that contained data in the column span.
    """
    left_col, right_col = col_span
    blank_streak = 0
    last_seen = start_row

    for r in range(start_row, len(grid) + 1):
        span = grid[r - 1][left_col - 1 : right_col]
        has_data = any(v is not None for v in span)

        if has_data:
            last_seen = r
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= max_blank_streak:
                break

    return last_seen


def slice_block_to_df(
    grid: List[List[object]],
    top: int,
    left: int,
    bottom: int,
    right: int,
    header_rows: int = 1,
    header_sep: str = " | ",
) -> pd.DataFrame:
    """Slice a rectangular block and convert it into a DataFrame.

    Supports 0..N header rows:
      - 0: auto-generate column names col_0..col_n
      - 1..N: flatten header rows using build_flat_columns()

    Args:
        grid: 2D grid of worksheet values.
        top: 1-based top row index (inclusive).
        left: 1-based left column index (inclusive).
        bottom: 1-based bottom row index (inclusive).
        right: 1-based right column index (inclusive).
        header_rows: Number of header rows at the top of the block.
        header_sep: Separator used when joining multi-row header parts.

    Returns:
        A pandas DataFrame containing the sliced table.
    """
    if top < 1 or left < 1 or bottom < top or right < left:
        return pd.DataFrame()

    block = [row[left - 1 : right] for row in grid[top - 1 : bottom]]
    if not block:
        return pd.DataFrame()

    header_rows = max(0, min(header_rows, len(block)))
    header_part = block[:header_rows]
    body = block[header_rows:]

    if header_rows == 0:
        cols = [f"col_{i}" for i in range(len(block[0]))]
    else:
        cols = build_flat_columns(header_part, sep=header_sep)

    df = pd.DataFrame(body, columns=cols)
    df = df.dropna(how="all").dropna(axis=1, how="all")
    return df


def extract_date_from_row(
    grid: List[List[object]],
    row: int,
    *,
    formats: Sequence[str] = ("%d-%b-%y",),
    strict: bool = False,
) -> Optional[str]:
    """Extract the first date found in a given grid row.

    Handles:
      - Real Excel date/datetime values (openpyxl returns date/datetime objects)
      - Exact strings that match one of the provided formats (e.g. "22-Jan-26")
      - If strict=False, date-like substrings inside longer text

    Args:
        grid: 2D grid of worksheet values.
        row: 1-based row index to search.
        formats: Accepted date formats (datetime.strptime). First format is also output format.
        strict: If True, only accept full-cell matches. If False, also search substrings.

    Returns:
        A date formatted with formats[0], or None if no date is found.
    """
    if row < 1 or row > len(grid):
        return None

    target_fmt = formats[0]
    cells = grid[row - 1]

    def normalise(s: str) -> str:
        return s.replace("\u00a0", " ").strip()

    for v in cells:
        if isinstance(v, (dt.date, dt.datetime)):
            d = v.date() if isinstance(v, dt.datetime) else v
            return d.strftime(target_fmt)

    for v in cells:
        if isinstance(v, str):
            s = normalise(v)
            for fmt in formats:
                try:
                    parsed = dt.datetime.strptime(s, fmt).date()
                    return parsed.strftime(target_fmt)
                except ValueError:
                    continue

    if strict:
        return None

    date_token_re = re.compile(r"\b(\d{1,2}-[A-Za-z]{3}-\d{2,4})\b")

    for v in cells:
        if not isinstance(v, str):
            continue
        s = normalise(v)
        m = date_token_re.search(s)
        if not m:
            continue

        token = m.group(1)
        parts = token.split("-")
        if len(parts) == 3:
            token = f"{parts[0]}-{parts[1].title()}-{parts[2]}"

        for fmt in formats:
            try:
                parsed = dt.datetime.strptime(token, fmt).date()
                return parsed.strftime(target_fmt)
            except ValueError:
                continue

    return None


@dataclass(frozen=True)
class TableSpec:
    """Specification for extracting a table from a worksheet grid.

    Attributes:
        name: Identifier for the extracted table.
        start_row: 1-based row index where the table block starts (header begins here).
        left_col: 1-based left column index (inclusive).
        right_col: 1-based right column index (inclusive).
        header_rows: Number of header rows (0..N) to interpret.
        end_row: If provided, use as the table end row. If None, infer the end.
        max_blank_streak: Only used when end_row is None. Ends the scan after N blank rows.
    """
    name: str
    start_row: int
    left_col: int
    right_col: int
    header_rows: int = 1
    end_row: Optional[int] = None
    max_blank_streak: int = 1


def parse_tables_with_specs(
    xlsx_path: str,
    sheet_name: str | int = 0,
    specs: Sequence[TableSpec] = (),
    header_sep: str = " | ",
) -> Dict[str, pd.DataFrame]:
    """Parse multiple tables from a single worksheet using explicit table specs.

    This is designed for report-like Excel sheets where multiple tables exist on one tab,
    often with merged header cells and 1..N header rows per table.

    Steps:
      1) Load workbook with openpyxl (preserves merges and cell types).
      2) Fill merged cells to normalise headers.
      3) Convert worksheet values into a 2D grid.
      4) For each TableSpec, slice a rectangular region into a DataFrame.
         If end_row is None, infer table end based on last row with data.

    Args:
        xlsx_path: Path to the .xlsx file.
        sheet_name: Worksheet name or 0-based sheet index.
        specs: List of TableSpec describing each table.
        header_sep: Separator used when joining multi-row header parts.

    Returns:
        A dict mapping spec.name to the extracted pandas DataFrame.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    normalise_merged_cells(ws)
    grid = sheet_to_grid(ws)

    out: Dict[str, pd.DataFrame] = {}
    for spec in specs:
        end = spec.end_row
        if end is None:
            end = last_data_row_in_span(
                grid=grid,
                start_row=spec.start_row,
                col_span=(spec.left_col, spec.right_col),
                max_blank_streak=spec.max_blank_streak,
            )

        out[spec.name] = slice_block_to_df(
            grid=grid,
            top=spec.start_row,
            left=spec.left_col,
            bottom=end,
            right=spec.right_col,
            header_rows=spec.header_rows,
            header_sep=header_sep,
        )

    return out

def excel_col_to_num(col: str) -> int:
    """
    Convert an Excel column label (e.g., 'A', 'FX') to a 1-based column number.

    Args:
        col: Excel column label.

    Returns:
        1-based column index (A -> 1, Z -> 26, AA -> 27).

    Raises:
        ValueError: If `col` is empty or contains non A-Z letters.
    """
    s = col.strip().upper()
    if not s or any(not ("A" <= ch <= "Z") for ch in s):
        raise ValueError(f"Invalid Excel column: {col!r}")

    num = 0
    for ch in s:
        num = num * 26 + (ord(ch) - ord("A") + 1)
    return num

# ----------------------------
# Example usage
# ----------------------------
# specs = [
#     TableSpec(name="table_a", start_row=2, left_col=1, right_col=8, header_rows=3, end_row=25),
#     TableSpec(name="table_b", start_row=30, left_col=2, right_col=10, header_rows=1, end_row=None, max_blank_streak=2),
#     TableSpec(name="table_c", start_row=80, left_col=1, right_col=6, header_rows=2, end_row=120),
# ]
#
# tables = parse_tables_with_specs("input.xlsx", sheet_name="Sheet1", specs=specs)
# date_str = extract_date_from_row(sheet_to_grid(load_workbook("input.xlsx", data_only=True)["Sheet1"]), 35)
#
# tables["table_a"].to_csv("table_a.csv", index=False)